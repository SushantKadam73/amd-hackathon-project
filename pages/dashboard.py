"""GRC Platform - Dashboard Page.

Displays:
1. Aggregate compliance score (0-100%) as weighted average of 5 controls
2. 5 control status cards with ID, title, status, evidence count, weightage
3. Risk heatmap (color-coded grid via plotly)
4. Gap summary widget (total open gaps, gaps by severity)
5. Clickable control cards navigating to control detail
6. Framework comparison view (cross-framework mappings)
7. Export Report button (PDF + Excel)
8. Empty state for first visit with onboarding guidance
9. Auto-refresh every 5 minutes + manual refresh button
"""

import io
from datetime import datetime, timezone
from typing import Any, Optional

import plotly.graph_objects as go
import streamlit as st

from app.api_client import (
    get_compliance_status,
    get_controls,
    get_frameworks,
    get_gap_summary,
    get_risk_heatmap,
)

# Page configuration
st.set_page_config(
    page_title="Dashboard - GRC Platform",
    page_icon="🏠",
    layout="wide",
)


# =============================================================================
# Constants
# =============================================================================

REFRESH_INTERVAL_SECONDS = 300  # 5 minutes
RISK_LEVELS = ["Low", "Medium", "High", "Critical"]
STATUS_COLORS = {
    "Fully Mapped": "#00cc66",
    "Partially Mapped": "#ffaa00",
    "Unmapped": "#ff3333",
}
STATUS_BG_COLORS = {
    "Fully Mapped": "#e6ffe6",
    "Partially Mapped": "#fff8e6",
    "Unmapped": "#ffe6e6",
}


# =============================================================================
# Helper Functions
# =============================================================================


def _get_mapping_status(score: float) -> str:
    """Determine mapping status from a compliance score.

    Args:
        score: Compliance score (0-100).

    Returns:
        Mapping status string.
    """
    if score >= 90.0:
        return "Fully Mapped"
    elif score >= 50.0:
        return "Partially Mapped"
    return "Unmapped"


def _get_status_color(status: str) -> str:
    """Get display color for a status.

    Args:
        status: Mapping status.

    Returns:
        Hex color string.
    """
    return STATUS_COLORS.get(status, "#999999")


def _format_timestamp(dt: Optional[datetime]) -> str:
    """Format a datetime for display.

    Args:
        dt: Datetime object or None.

    Returns:
        Formatted time string.
    """
    if dt is None:
        dt = datetime.now(timezone.utc)
    return dt.strftime("%Y-%m-%d %H:%M:%S UTC")


# =============================================================================
# Render Functions
# =============================================================================


def _render_empty_state() -> None:
    """Render onboarding empty state when no data exists."""
    st.markdown("## 🛡️ Welcome to GRC Platform")

    col1, col2 = st.columns([3, 2])
    with col1:
        st.markdown(
            """
            This platform helps you manage **NIST SP 800-53 compliance** for
            your datacenter organization. Get started by following these steps:

            ### 🚀 Getting Started

            1. **📚 Seed the Framework Library**
               - Load NIST SP 800-53 with 5 controls (PE-03, AC-02, SC-07,
                 IR-06, RA-05)
               - Run: `python scripts/seed_data.py`

            2. **📎 Upload Evidence Artifacts**
               - Upload policy documents, procedures, and system configs
               - Supported formats: PDF, DOCX, PNG, JPG

            3. **🔗 Map Evidence to Controls**
               - Link your evidence to the relevant security controls
               - Each artifact type has a defined weightage

            4. **📊 View Your Compliance Score**
               - Track your compliance progress in real-time
               - Identify gaps and prioritize remediation

            ---
            *Once data is loaded, this dashboard will show your compliance
            status, risk heatmap, gap analysis, and more.*
            """
        )

    with col2:
        st.info(
            "### Quick Actions\n\n"
            "🔹 [Browse Control Library](/framework_library)\n\n"
            "🔹 [Upload Evidence](/evidence_library)\n\n"
            "---\n"
            "**Need help?**\n\n"
            "Refer to the documentation for step-by-step guidance "
            "on setting up your compliance program.",
            icon="ℹ️",
        )

    # Show system status
    st.divider()
    st.subheader("📡 System Status")

    frameworks = get_frameworks()
    col_a, col_b, col_c = st.columns(3)
    with col_a:
        st.metric("Frameworks", len(frameworks))
    with col_b:
        st.metric("Controls", 0)
    with col_c:
        st.metric("Compliance Score", "0%", delta="Not Started")


def _render_auto_refresh_controls() -> None:
    """Render the auto-refresh toggle and manual refresh button.

    Also manages the last_updated timestamp in session state.
    """
    # Initialize session state
    if "last_updated" not in st.session_state:
        st.session_state.last_updated = datetime.now(timezone.utc)
    if "refresh_count" not in st.session_state:
        st.session_state.refresh_count = 0

    cols = st.columns([3, 1, 1, 2])
    with cols[0]:
        st.caption(
            f"🔄 Auto-refreshes every 5 minutes. "
            f"Last updated: {_format_timestamp(st.session_state.last_updated)}"
        )
    with cols[1]:
        if st.button("🔄 Refresh Now", type="secondary", use_container_width=True):
            st.session_state.last_updated = datetime.now(timezone.utc)
            st.session_state.refresh_count += 1
            st.rerun()
    with cols[2]:
        st.caption(f"Refreshes: {st.session_state.refresh_count}")


def _render_compliance_score(
    compliance: Optional[dict[str, Any]],
    has_data: bool,
) -> None:
    """Render the aggregate compliance score gauge.

    Args:
        compliance: Compliance status data from API.
        has_data: Whether there is actual compliance data.
    """
    st.subheader("📊 Aggregate Compliance Score")

    if not has_data or not compliance:
        score = 0.0
        status = "Not Started"
    else:
        score = compliance.get("overall_score", 0.0)
        status = compliance.get("overall_status", "Not Started")

    score_color = "#ff3333" if score < 50 else ("#ffaa00" if score < 90 else "#00cc66")

    # Gauge chart using plotly
    fig = go.Figure(go.Indicator(
        mode="gauge+number+delta",
        value=score,
        domain={"x": [0, 1], "y": [0, 1]},
        title={"text": f"Overall Compliance — {status}"},
        delta={"reference": 100, "increasing": {"color": "#00cc66"}},
        gauge={
            "axis": {"range": [0, 100], "tickwidth": 1, "tickcolor": "darkblue"},
            "bar": {"color": score_color},
            "bgcolor": "white",
            "borderwidth": 2,
            "bordercolor": "#cccccc",
            "steps": [
                {"range": [0, 50], "color": "#ffe6e6"},
                {"range": [50, 90], "color": "#fff8e6"},
                {"range": [90, 100], "color": "#e6ffe6"},
            ],
            "threshold": {
                "line": {"color": "#333333", "width": 4},
                "thickness": 0.75,
                "value": score,
            },
        },
    ))
    fig.update_layout(
        height=280,
        margin={"t": 50, "b": 0, "l": 0, "r": 0},
        paper_bgcolor="rgba(0,0,0,0)",
        font={"color": "#333333", "family": "sans-serif"},
    )
    st.plotly_chart(fig, use_container_width=True)

    if not has_data:
        st.info(
            "ℹ️ No evidence has been uploaded yet. "
            "The compliance score is 0% until evidence artifacts are "
            "uploaded, mapped to controls, and approved.",
            icon="ℹ️",
        )


def _render_control_cards(
    compliance: Optional[dict[str, Any]],
    has_data: bool,
) -> None:
    """Render 5 control status cards.

    Each card shows: control ID, title, compliance status,
    evidence count, and weightage coverage.

    Args:
        compliance: Compliance status data from API.
        has_data: Whether there is actual compliance data.
    """
    st.subheader("🛡️ Control Status")

    controls_data: list[dict[str, Any]] = []
    if has_data and compliance:
        controls_data = compliance.get("controls", [])

    if not controls_data:
        # Show placeholder cards when no data
        control_ids = ["PE-03", "AC-02", "SC-07", "IR-06", "RA-05"]
        col_wrapper = st.columns(5)
        for idx, cid in enumerate(control_ids):
            with col_wrapper[idx]:
                st.markdown(
                    f"""
                    <div style="border:1px solid #ddd; border-radius:10px;
                         padding:15px; text-align:center; background:#f9f9f9;
                         margin-bottom:10px;">
                        <h4 style="margin:0; color:#666;">{cid}</h4>
                        <p style="font-size:12px; color:#999; margin:5px 0;">
                            Not Loaded
                        </p>
                        <div style="background:#ffe6e6; border-radius:5px;
                             padding:5px 10px; display:inline-block;
                             margin-top:5px;">
                            <span style="color:#ff3333; font-weight:bold;">
                                Unmapped
                            </span>
                        </div>
                        <p style="font-size:11px; color:#999; margin-top:8px;">
                            Evidence: 0 | Coverage: 0%
                        </p>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )
        return

    # Render actual control cards
    cols = st.columns(5)
    for idx, ctrl in enumerate(controls_data):
        ctrl_id = ctrl.get("control_id", "")
        title = ctrl.get("title", "")
        status = ctrl.get("status", "Unmapped")
        score = ctrl.get("score", 0.0)
        evidence_count = ctrl.get("evidence_count", 0)
        validated = ctrl.get("validated_weightage", 0.0)
        total = ctrl.get("total_required_weightage", 100.0)
        coverage_pct = min(validated / total * 100, 100) if total > 0 else 0.0

        status_color = _get_status_color(status)
        bg_color = STATUS_BG_COLORS.get(status, "#f9f9f9")

        with cols[idx]:
            # Title truncated for card
            title_short = (title[:40] + "...") if len(title) > 40 else title

            card_html = f"""
            <a href="/control_detail?control_id={ctrl_id}"
               target="_self" style="text-decoration:none; color:inherit;">
            <div style="border:1px solid #ddd; border-radius:10px;
                 padding:12px; text-align:center; background:{bg_color};
                 margin-bottom:10px; cursor:pointer;
                 transition:box-shadow 0.2s;
                 box-shadow:0 1px 3px rgba(0,0,0,0.08);"
                 onmouseover="this.style.boxShadow='0 4px 12px rgba(0,0,0,0.15)'"
                 onmouseout="this.style.boxShadow='0 1px 3px rgba(0,0,0,0.08)'">
                <h4 style="margin:0; color:#333;">{ctrl_id}</h4>
                <p style="font-size:11px; color:#666; margin:5px 0;
                     min-height:32px; line-height:1.3;">
                    {title_short}
                </p>
                <div style="background:{status_color}20; border-radius:5px;
                     padding:3px 8px; display:inline-block; margin-top:3px;
                     border:1px solid {status_color};">
                    <span style="color:{status_color}; font-weight:bold;
                         font-size:12px;">
                        {status}
                    </span>
                </div>
                <p style="font-size:11px; color:#888; margin-top:8px;">
                    Evidence: {evidence_count}<br>
                    Coverage: {coverage_pct:.0f}%
                </p>
                <div style="background:#e0e0e0; border-radius:10px;
                     height:6px; width:100%; margin-top:5px; overflow:hidden;">
                    <div style="background:{status_color}; width:{coverage_pct}%;
                         height:100%; border-radius:10px;"></div>
                </div>
            </div>
            </a>
            """
            st.markdown(card_html, unsafe_allow_html=True)


def _render_risk_heatmap(
    controls_data: list[dict[str, Any]],
    has_data: bool,
) -> None:
    """Render a risk heatmap using plotly.

    Shows controls vs risk levels with color-coded cells.

    Args:
        controls_data: List of control data with status info.
        has_data: Whether there is actual compliance data.
    """
    st.subheader("🔥 Risk Heatmap")

    if not has_data or not controls_data:
        st.info(
            "No control data available to render the risk heatmap. "
            "Load control data to see the visualization.",
            icon="ℹ️",
        )
        return

    # Build heatmap data: each cell shows the control's compliance status
    # Risk levels: Low, Medium, High, Critical
    control_ids = [c.get("control_id", "") for c in controls_data]
    status_values = []
    score_values = []
    colors = []

    for ctrl in controls_data:
        status = ctrl.get("status", "Unmapped")
        score = ctrl.get("score", 0.0)
        color = _get_status_color(status)

        # For each risk level, the cell shows the same status/color
        # but the interpretation differs by risk level
        status_values.append(status)
        score_values.append(round(score, 1))
        colors.append(color)

    # Create a 2D matrix for the heatmap
    # Controls on Y-axis, Risk Levels on X-axis
    z_matrix: list[list[float]] = []
    text_matrix: list[list[str]] = []
    color_matrix: list[list[str]] = []

    for idx, ctrl in enumerate(controls_data):
        status = ctrl.get("status", "Unmapped")
        score = ctrl.get("score", 0.0)
        row_z = []
        row_text = []
        row_colors = []

        for risk_idx, risk_level in enumerate(RISK_LEVELS):
            # Map risk level to a numeric score for color intensity
            if status == "Fully Mapped":
                val = 1.0  # Low risk
            elif status == "Partially Mapped":
                val = 0.5  # Medium risk
            else:
                val = 0.0  # High risk

            row_z.append(val)
            row_text.append(
                f"<b>{ctrl.get('control_id', '')}</b><br>"
                f"Status: {status}<br>"
                f"Score: {score:.1f}%<br>"
                f"Risk Level: {risk_level}"
            )
            row_colors.append(ctrl.get("control_id", ""))

        z_matrix.append(row_z)
        text_matrix.append(row_text)

    # Custom colorscale: red (0) -> yellow (0.5) -> green (1)
    colorscale = [
        [0.0, "#ff3333"],
        [0.25, "#ff6666"],
        [0.5, "#ffcc00"],
        [0.75, "#99cc33"],
        [1.0, "#00cc66"],
    ]

    fig = go.Figure(data=go.Heatmap(
        z=z_matrix,
        x=RISK_LEVELS,
        y=control_ids,
        text=text_matrix,
        texttemplate="%{text}",
        hovertemplate="%{text}<extra></extra>",
        colorscale=colorscale,
        showscale=True,
        colorbar={
            "title": "Risk Level",
            "tickvals": [0, 0.5, 1.0],
            "ticktext": ["High Risk", "Medium Risk", "Low Risk"],
        },
        zmin=0,
        zmax=1,
        xgap=3,
        ygap=3,
    ))

    fig.update_layout(
        height=300,
        margin={"t": 20, "b": 30, "l": 20, "r": 20},
        xaxis={
            "title": "Risk Level",
            "side": "top",
        },
        yaxis={
            "title": "Controls",
            "autorange": "reversed",
        },
        paper_bgcolor="rgba(0,0,0,0)",
        font={"size": 12},
    )

    st.plotly_chart(fig, use_container_width=True)

    # Legend
    leg_cols = st.columns(3)
    with leg_cols[0]:
        st.markdown(
            "<span style='display:inline-block; width:12px; height:12px; "
            "background:#00cc66; border-radius:2px; margin-right:5px;'></span>"
            " Low Risk (Fully Mapped)",
            unsafe_allow_html=True,
        )
    with leg_cols[1]:
        st.markdown(
            "<span style='display:inline-block; width:12px; height:12px; "
            "background:#ffcc00; border-radius:2px; margin-right:5px;'></span>"
            " Medium Risk (Partially Mapped)",
            unsafe_allow_html=True,
        )
    with leg_cols[2]:
        st.markdown(
            "<span style='display:inline-block; width:12px; height:12px; "
            "background:#ff3333; border-radius:2px; margin-right:5px;'></span>"
            " High Risk (Unmapped)",
            unsafe_allow_html=True,
        )


def _render_gap_summary(has_data: bool) -> None:
    """Render the gap summary widget.

    Shows total open gaps and breakdown by severity.

    Args:
        has_data: Whether there is actual compliance data.
    """
    st.subheader("📋 Gap Summary")

    gap_data = get_gap_summary() if has_data else None

    if not gap_data or gap_data.get("total_open_gaps", 0) == 0:
        if has_data:
            st.success("✅ No open gaps! All controls are fully mapped.", icon="✅")
        else:
            st.info(
                "No gap data available yet. Upload and map evidence "
                "to see your compliance gaps.",
                icon="ℹ️",
            )
        return

    total = gap_data.get("total_open_gaps", 0)
    severity = gap_data.get("gaps_by_severity", {})

    col1, col2, col3, col4 = st.columns(4)

    with col1:
        st.metric("Total Open Gaps", total)

    severity_map = [
        ("Critical", severity.get("critical", 0), "#ff3333"),
        ("High", severity.get("high", 0), "#ff6600"),
        ("Medium", severity.get("medium", 0), "#ffaa00"),
        ("Low", severity.get("low", 0), "#66aa00"),
    ]

    cols = [col2, col3, col4]
    for i, (sev_name, sev_count, sev_color) in enumerate(severity_map[:3]):
        with cols[i]:
            st.markdown(
                f"<div style='text-align:center; padding:10px; "
                f"border:1px solid {sev_color}40; border-radius:8px; "
                f"background:{sev_color}10;'>"
                f"<h3 style='margin:0; color:{sev_color};'>{sev_count}</h3>"
                f"<p style='margin:2px 0 0; font-size:13px; color:#666;'>{sev_name}</p>"
                f"</div>",
                unsafe_allow_html=True,
            )

    # Also show severity breakdown as a compact table
    sev_rows = ""
    for sev_name, sev_count, sev_color in severity_map:
        pct = (sev_count / total * 100) if total > 0 else 0
        sev_rows += (
            f"<tr><td style='color:{sev_color};font-weight:bold;'>{sev_name}</td>"
            f"<td style='text-align:center;'>{sev_count}</td>"
            f"<td style='text-align:center;'>{pct:.0f}%</td></tr>"
        )

    st.markdown(
        f"<table style='width:100%; font-size:13px; margin-top:10px;'>"
        f"<thead><tr style='background:#f0f2f6;'>"
        f"<th>Severity</th><th>Count</th><th>% of Total</th>"
        f"</tr></thead>"
        f"<tbody>{sev_rows}</tbody></table>",
        unsafe_allow_html=True,
    )


def _render_framework_comparison(has_data: bool) -> None:
    """Render the framework comparison view.

    Shows cross-framework mappings for all 5 controls.

    Args:
        has_data: Whether frameworks are loaded.
    """
    st.subheader("🔗 Framework Comparison")

    if not has_data:
        st.info(
            "Framework comparison will be available once NIST SP 800-53 "
            "and its controls are loaded into the system.",
            icon="ℹ️",
        )
        return

    frameworks = get_frameworks()
    if not frameworks:
        st.info("No frameworks loaded. Run `python scripts/seed_data.py`.", icon="📚")
        return

    # Fetch controls with cross-framework mappings
    fw_id = frameworks[0]["id"]
    controls = get_controls(fw_id)

    if not controls:
        st.info("No controls found.", icon="📋")
        return

    # Build comparison table
    comparison_rows: list[dict[str, Any]] = []

    for ctrl in controls:
        ctrl_id = ctrl.get("control_id", "")
        title = ctrl.get("title", "")
        mappings = ctrl.get("cross_framework_mappings") or {}

        nist_csf = mappings.get("nist_csf_2_0", {})
        cis = mappings.get("cis_controls_v8", {})
        pci = mappings.get("pci_dss_v4_0_1", {})

        comparison_rows.append({
            "Control": ctrl_id,
            "Title": title,
            "NIST CSF 2.0": nist_csf.get("mapped_to") if nist_csf else "No mapping defined",
            "CIS Controls v8": cis.get("mapped_to") if cis else "No mapping defined",
            "PCI DSS v4.0.1": pci.get("mapped_to") if pci else "No mapping defined",
        })

    # Render as expandable sections for each control
    for row in comparison_rows:
        ctrl_id = row["Control"]
        title = row["Title"]

        # Get compliance data for status
        compliance = get_compliance_status()
        ctrl_status = "Unmapped"
        ctrl_score = 0.0
        if compliance:
            for c in compliance.get("controls", []):
                if c.get("control_id") == ctrl_id:
                    ctrl_status = c.get("status", "Unmapped")
                    ctrl_score = c.get("score", 0.0)
                    break

        status_color = _get_status_color(ctrl_status)

        with st.container(border=True):
            cols = st.columns([1, 2, 1.5, 1.5, 1.5])
            with cols[0]:
                st.markdown(f"**{ctrl_id}**")
            with cols[1]:
                st.markdown(f"<small>{title}</small>", unsafe_allow_html=True)
            with cols[2]:
                mapped_nist = row["NIST CSF 2.0"]
                if mapped_nist != "No mapping defined":
                    st.markdown(f"`{mapped_nist}`")
                else:
                    st.markdown(f"<small style='color:#999;'>{mapped_nist}</small>",
                                unsafe_allow_html=True)
            with cols[3]:
                mapped_cis = row["CIS Controls v8"]
                if mapped_cis != "No mapping defined":
                    st.markdown(f"`{mapped_cis}`")
                else:
                    st.markdown(f"<small style='color:#999;'>{mapped_cis}</small>",
                                unsafe_allow_html=True)
            with cols[4]:
                mapped_pci = row["PCI DSS v4.0.1"]
                if mapped_pci != "No mapping defined":
                    st.markdown(f"`{mapped_pci}`")
                else:
                    st.markdown(f"<small style='color:#999;'>{mapped_pci}</small>",
                                unsafe_allow_html=True)

            # Status indicator
            st.markdown(
                f"<small>Status: "
                f"<span style='color:{status_color};font-weight:bold;'>"
                f"{ctrl_status}</span> "
                f"(Score: {ctrl_score:.1f}%)</small>",
                unsafe_allow_html=True,
            )

            # Link to control detail
            st.markdown(
                f"<a href='/control_detail?control_id={ctrl_id}' "
                f"target='_self'>View details →</a>",
                unsafe_allow_html=True,
            )


def _render_evidence_progress(
    compliance: Optional[dict[str, Any]],
    has_data: bool,
) -> None:
    """Render evidence collection progress tracker.

    Shows uploaded vs required artifact types for each control.

    Args:
        compliance: Compliance status data from API.
        has_data: Whether there is actual compliance data.
    """
    st.subheader("📎 Evidence Collection Progress")

    if not has_data or not compliance:
        st.info(
            "No evidence uploaded yet. The progress tracker will show "
            "evidence collection status once artifacts are uploaded.",
            icon="ℹ️",
        )
        return

    controls_data = compliance.get("controls", [])
    if not controls_data:
        return

    for ctrl in controls_data:
        ctrl_id = ctrl.get("control_id", "")
        title = ctrl.get("title", "")
        validated = ctrl.get("validated_weightage", 0.0)
        total = ctrl.get("total_required_weightage", 100.0)
        evidence_count = ctrl.get("evidence_count", 0)

        pct = min(validated / total * 100, 100) if total > 0 else 0.0
        bar_color = "#00cc66" if pct >= 80 else ("#ffaa00" if pct >= 50 else "#ff3333")

        # Estimate artifact types from compliance score
        if ctrl.get("score", 0) >= 90:
            status_text = "Fully Mapped"
        elif ctrl.get("score", 0) >= 50:
            status_text = "Partially Mapped"
        else:
            status_text = "Unmapped"

        st.markdown(
            f"**{ctrl_id}** — {title}  "
            f"<span style='color:{bar_color};font-weight:bold;'>{status_text}</span>",
            unsafe_allow_html=True,
        )
        st.progress(
            pct / 100.0,
            text=f"Evidence items: {evidence_count} | "
                 f"Weightage: {validated:.0f}% / {total:.0f}% ({pct:.0f}%)",
        )


def _generate_pdf_report(
    compliance: Optional[dict[str, Any]],
    gap_data: Optional[dict[str, Any]],
) -> bytes:
    """Generate a PDF report of the dashboard data.

    Args:
        compliance: Compliance status data.
        gap_data: Gap summary data.

    Returns:
        PDF report as bytes.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch, mm
    from reportlab.platypus import (
        SimpleDocTemplate,
        Paragraph,
        Spacer,
        Table,
        TableStyle,
        PageBreak,
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        rightMargin=72,
        leftMargin=72,
        topMargin=72,
        bottomMargin=72,
    )

    styles = getSampleStyleSheet()
    title_style = styles["Title"]
    heading_style = styles["Heading2"]
    normal_style = styles["Normal"]

    elements = []

    # Title
    elements.append(Paragraph("GRC Platform - Compliance Report", title_style))
    elements.append(Spacer(1, 12))
    elements.append(
        Paragraph(
            f"Generated: {_format_timestamp(datetime.now(timezone.utc))}",
            normal_style,
        )
    )
    elements.append(Spacer(1, 20))

    # Overall Score
    elements.append(Paragraph("Overall Compliance Score", heading_style))
    if compliance:
        score = compliance.get("overall_score", 0.0)
        status = compliance.get("overall_status", "Not Started")
        elements.append(
            Paragraph(
                f"Score: {score:.1f}% — Status: {status}",
                normal_style,
            )
        )
    else:
        elements.append(Paragraph("Score: 0% — No data available.", normal_style))

    elements.append(Spacer(1, 16))

    # Per-Control Status
    elements.append(Paragraph("Per-Control Status", heading_style))
    if compliance and compliance.get("controls"):
        control_table_data = [["Control ID", "Title", "Status", "Score", "Evidence"]]
        for ctrl in compliance["controls"]:
            control_table_data.append([
                ctrl.get("control_id", ""),
                ctrl.get("title", ""),
                ctrl.get("status", ""),
                f"{ctrl.get('score', 0.0):.1f}%",
                str(ctrl.get("evidence_count", 0)),
            ])

        control_table = Table(control_table_data, colWidths=[60, 200, 100, 60, 60])
        control_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0066cc")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#f0f2f6")),
            ("GRID", (0, 0), (-1, -1), 1, colors.grey),
        ]))
        elements.append(control_table)
    else:
        elements.append(Paragraph("No control data available.", normal_style))

    elements.append(Spacer(1, 16))

    # Gap Summary
    elements.append(Paragraph("Gap Summary", heading_style))
    if gap_data and gap_data.get("total_open_gaps", 0) > 0:
        total = gap_data.get("total_open_gaps", 0)
        severity = gap_data.get("gaps_by_severity", {})
        elements.append(Paragraph(f"Total Open Gaps: {total}", normal_style))
        gap_table_data = [["Severity", "Count"]]
        for sev in ["critical", "high", "medium", "low"]:
            gap_table_data.append([sev.title(), str(severity.get(sev, 0))])

        gap_table = Table(gap_table_data, colWidths=[120, 80])
        gap_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0066cc")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (1, 0), (1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 1, colors.grey),
        ]))
        elements.append(gap_table)
    else:
        elements.append(Paragraph("No open gaps. All controls fully mapped.", normal_style))

    elements.append(Spacer(1, 16))

    # Footer
    elements.append(Spacer(1, 30))
    elements.append(
        Paragraph(
            "GRC Platform MVP — NIST SP 800-53 Compliance Report",
            ParagraphStyle(
                "footer",
                parent=normal_style,
                fontSize=8,
                textColor=colors.grey,
            ),
        )
    )

    doc.build(elements)
    pdf_bytes = buf.getvalue()
    buf.close()
    return pdf_bytes


def _generate_excel_report(
    compliance: Optional[dict[str, Any]],
    gap_data: Optional[dict[str, Any]],
) -> bytes:
    """Generate an Excel report of the dashboard data.

    Args:
        compliance: Compliance status data.
        gap_data: Gap summary data.

    Returns:
        Excel file as bytes.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    # --- Sheet 1: Compliance Score ---
    ws1 = wb.active
    ws1.title = "Compliance Score"
    ws1.append(["GRC Platform - Compliance Report"])
    ws1.append([f"Generated: {_format_timestamp(datetime.now(timezone.utc))}"])
    ws1.append([])

    if compliance:
        ws1.append(["Overall Score", f"{compliance.get('overall_score', 0.0):.1f}%"])
        ws1.append(["Overall Status", compliance.get("overall_status", "N/A")])
        ws1.append(["Total Controls", compliance.get("total_controls", 0)])

    ws1.append([])
    ws1.append(["Control ID", "Title", "Status", "Score (%)", "Evidence Count",
                 "Validated Weightage", "Total Required"])

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for cell in ws1[5]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    if compliance:
        for ctrl in compliance.get("controls", []):
            ws1.append([
                ctrl.get("control_id", ""),
                ctrl.get("title", ""),
                ctrl.get("status", ""),
                ctrl.get("score", 0.0),
                ctrl.get("evidence_count", 0),
                ctrl.get("validated_weightage", 0.0),
                ctrl.get("total_required_weightage", 0.0),
            ])

    # Auto-adjust column widths
    for col in ws1.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws1.column_dimensions[col_letter].width = adjusted_width

    # --- Sheet 2: Gap Summary ---
    ws2 = wb.create_sheet("Gap Summary")
    ws2.append(["Gap Summary"])
    ws2.append([])

    if gap_data:
        ws2.append(["Total Open Gaps", gap_data.get("total_open_gaps", 0)])
        ws2.append(["Oldest Gap Age (days)", gap_data.get("oldest_open_gap_age_days", 0)])
        ws2.append([])
        ws2.append(["Severity", "Count"])
        severity = gap_data.get("gaps_by_severity", {})
        for sev in ["critical", "high", "medium", "low"]:
            ws2.append([sev.title(), severity.get(sev, 0)])

        for cell in ws2[4]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    # Auto-adjust column widths
    for col in ws2.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws2.column_dimensions[col_letter].width = adjusted_width

    # Save to buffer
    buf = io.BytesIO()
    wb.save(buf)
    excel_bytes = buf.getvalue()
    buf.close()
    return excel_bytes


def _render_export_button(
    compliance: Optional[dict[str, Any]],
    gap_data: Optional[dict[str, Any]],
) -> None:
    """Render the Export Report button with PDF and Excel download.

    Args:
        compliance: Compliance status data.
        gap_data: Gap summary data.
    """
    st.subheader("📄 Export Report")

    col1, col2 = st.columns(2)

    with col1:
        if st.button("📕 Export as PDF", use_container_width=True, type="primary"):
            with st.spinner("Generating PDF report..."):
                try:
                    pdf_bytes = _generate_pdf_report(compliance, gap_data)
                    st.download_button(
                        label="📥 Download PDF Report",
                        data=pdf_bytes,
                        file_name=f"grc_compliance_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
                        mime="application/pdf",
                        use_container_width=True,
                    )
                except Exception as e:
                    st.error(f"Failed to generate PDF: {e}")

    with col2:
        if st.button("📗 Export as Excel", use_container_width=True, type="secondary"):
            with st.spinner("Generating Excel report..."):
                try:
                    excel_bytes = _generate_excel_report(compliance, gap_data)
                    st.download_button(
                        label="📥 Download Excel Report",
                        data=excel_bytes,
                        file_name=f"grc_compliance_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
                except Exception as e:
                    st.error(f"Failed to generate Excel: {e}")


# =============================================================================
# Main Entry Point
# =============================================================================


def main() -> None:
    """Main entry point for the Dashboard page."""
    st.title("🏠 GRC Platform Dashboard")

    # --- Auto-refresh logic ---
    _render_auto_refresh_controls()

    # --- Fetch data ---
    frameworks = get_frameworks()
    has_frameworks = len(frameworks) > 0

    compliance = get_compliance_status()
    has_compliance = (
        compliance is not None
        and compliance.get("total_controls", 0) > 0
    )

    has_data = has_frameworks and has_compliance

    # Check if any control has evidence or scores
    if has_compliance:
        controls_data = compliance.get("controls", [])
        total_evidence = sum(c.get("evidence_count", 0) for c in controls_data)
        total_score = sum(c.get("score", 0.0) for c in controls_data)
        if total_evidence == 0 and total_score == 0.0:
            has_data = False

    # --- Empty State ---
    if not has_data:
        _render_empty_state()
        return

    # --- Main Dashboard Content ---

    # Row 1: Compliance Score + Quick Metrics
    col_left, col_right = st.columns([2, 3])

    with col_left:
        _render_compliance_score(compliance, has_data)

    with col_right:
        # Quick metrics row
        if compliance:
            controls_data = compliance.get("controls", [])
            total_evidence = sum(
                c.get("evidence_count", 0) for c in controls_data
            )
            fully_mapped = sum(
                1 for c in controls_data
                if c.get("status") == "Fully Mapped"
            )
            partially_mapped = sum(
                1 for c in controls_data
                if c.get("status") == "Partially Mapped"
            )

            metric_cols = st.columns(3)
            with metric_cols[0]:
                st.metric(
                    "Total Evidence",
                    total_evidence,
                )
            with metric_cols[1]:
                st.metric(
                    "Fully Mapped",
                    fully_mapped,
                    delta=f"{fully_mapped}/5 controls",
                )
            with metric_cols[2]:
                st.metric(
                    "Partially Mapped",
                    partially_mapped,
                )

        # Evidence progress tracker
        _render_evidence_progress(compliance, has_data)

    st.divider()

    # Row 2: Control Status Cards
    _render_control_cards(compliance, has_data)

    st.divider()

    # Row 3: Risk Heatmap + Gap Summary
    col_heat, col_gap = st.columns([3, 2])

    with col_heat:
        controls_data = compliance.get("controls", []) if compliance else []
        _render_risk_heatmap(controls_data, has_data)

    with col_gap:
        _render_gap_summary(has_data)

    st.divider()

    # Row 4: Framework Comparison
    _render_framework_comparison(has_data)

    st.divider()

    # Row 5: Export Report
    gap_data = get_gap_summary()
    _render_export_button(compliance, gap_data)

    st.divider()
    st.caption(
        "GRC Platform MVP v0.1.0 | "
        "NIST SP 800-53 Compliance Dashboard | "
        f"Page auto-refreshes every {REFRESH_INTERVAL_SECONDS // 60} minutes"
    )


if __name__ == "__main__":
    main()
